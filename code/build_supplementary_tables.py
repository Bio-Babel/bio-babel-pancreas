#!/usr/bin/env python3
"""Build the publication-facing supplementary-table workbook.

The workbook deliberately contains only results that support claims retained
in the manuscript or figures. More granular audit tables remain in ``tables/``.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
TABLES = ROOT / "tables"
OUTPUT = ROOT / "supplementary_tables.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_GREY = Side(style="thin", color="D9E1F2")


def read_tsv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t")


def samples_qc() -> pd.DataFrame:
    frame = read_tsv(ROOT / "samples.tsv").rename(
        columns={
            "sample": "sample_id",
            "gsm": "GEO_accession",
            "srr": "SRA_run",
            "genotype": "genotype_reported",
            "inferred_sex": "sex_inferred_from_Y_markers",
        }
    )
    genotype_short = {
        "control": "control",
        "Brg1Δendo;Brm+/-": "dEndo",
        "DKO": "DKO",
    }
    frame.insert(
        frame.columns.get_loc("genotype_reported") + 1,
        "genotype_short",
        frame["genotype_reported"].map(genotype_short),
    )
    order = pd.Categorical(
        frame["genotype_short"],
        categories=["control", "dEndo", "DKO"],
        ordered=True,
    )
    frame = (
        frame.assign(_genotype_order=order)
        .sort_values(["_genotype_order", "sample_id"])
        .drop(columns="_genotype_order")
        .reset_index(drop=True)
    )
    return frame


def beam_results() -> tuple[pd.DataFrame, pd.DataFrame]:
    beam = pd.read_csv(TABLES / "beam.csv.gz")
    displayed = read_tsv(TABLES / "fig2h.tsv")
    label_by_id = displayed.set_index("gene_id")["labelled"].to_dict()
    displayed_ids = set(displayed["gene_id"])

    beam = beam.copy()
    beam.insert(0, "BEAM_rank", np.arange(1, len(beam) + 1))
    beam["BEAM_q_lt_1e-4"] = beam["qval"] < 1e-4
    beam["included_in_top100_heatmap"] = beam["gene_id"].isin(displayed_ids)
    beam["labelled_in_top100_heatmap"] = (
        beam["gene_id"].map(label_by_id).eq(True)
    )
    beam = beam.rename(
        columns={
            "gene_short_name": "gene_symbol",
            "pval": "BEAM_p_value",
            "qval": "BEAM_BH_q_value",
            "highly_variable": "highly_variable_in_input",
            "use_for_ordering": "used_for_DDRTree_ordering",
        }
    )
    keep = [
        "BEAM_rank",
        "gene_id",
        "gene_symbol",
        "status",
        "BEAM_p_value",
        "BEAM_BH_q_value",
        "num_cells_expressed",
        "highly_variable_in_input",
        "used_for_DDRTree_ordering",
        "BEAM_q_lt_1e-4",
        "included_in_top100_heatmap",
        "labelled_in_top100_heatmap",
    ]
    return beam[keep].copy(), beam[["gene_id", "BEAM_rank"]].copy()


def genotype_screen(beam_ranks: pd.DataFrame) -> pd.DataFrame:
    screen = read_tsv(TABLES / "screen_control.tsv.gz")
    gene_meta = read_tsv(TABLES / "screen_genes.tsv.gz")
    display = read_tsv(TABLES / "sfig6h.tsv").rename(
        columns={"gene": "gene_symbol", "category": "SFig6h_category"}
    )
    display["contrast"] = display["contrast"].map(
        {
            "dEndo − control": "dEndo-control",
            "DKO − control": "DKO-control",
        }
    )
    display = display[["gene_symbol", "contrast", "SFig6h_category"]]

    measures = [
        "converged",
        "mean_signed_log2fc",
        "pvalue_l2fc_0_5",
        "qvalue_l2fc_0_5",
    ]
    alpha = screen.loc[screen["lineage"].eq("alpha"), [
        "gene_id", "gene_symbol", "contrast", *measures
    ]].rename(columns={name: f"alpha_{name}" for name in measures})
    beta = screen.loc[screen["lineage"].eq("beta"), [
        "gene_id", "gene_symbol", "contrast", *measures
    ]].rename(columns={name: f"beta_{name}" for name in measures})
    wide = alpha.merge(
        beta,
        on=["gene_id", "gene_symbol", "contrast"],
        how="inner",
        validate="one_to_one",
    )
    wide = wide.merge(
        gene_meta[[
            "gene_id",
            "num_cells_expressed",
            "technical_class",
            "original_marker",
        ]],
        on="gene_id",
        how="left",
        validate="many_to_one",
    ).merge(
        beam_ranks,
        on="gene_id",
        how="left",
        validate="many_to_one",
    ).merge(
        display,
        on=["gene_symbol", "contrast"],
        how="left",
        validate="one_to_one",
    )

    both_significant = (
        wide["alpha_qvalue_l2fc_0_5"].lt(0.05)
        & wide["beta_qvalue_l2fc_0_5"].lt(0.05)
    )
    both_positive = (
        wide["alpha_mean_signed_log2fc"].gt(0)
        & wide["beta_mean_signed_log2fc"].gt(0)
    )
    both_negative = (
        wide["alpha_mean_signed_log2fc"].lt(0)
        & wide["beta_mean_signed_log2fc"].lt(0)
    )
    wide["cross_lineage_result"] = np.select(
        [both_significant & both_positive, both_significant & both_negative],
        ["shared increase", "shared decrease"],
        default="not shared at the specified threshold",
    )
    wide["max_lineage_BH_q_value"] = wide[[
        "alpha_qvalue_l2fc_0_5", "beta_qvalue_l2fc_0_5"
    ]].max(axis=1)
    wide["passes_SFig6h_display_filter"] = wide["SFig6h_category"].notna()
    wide["SFig6h_category"] = wide["SFig6h_category"].fillna("not displayed")
    wide["highlighted_candidate"] = wide["gene_symbol"].eq("Kcnip4")
    wide["BEAM_top1000"] = wide["BEAM_rank"].le(1000)

    cross_contrast = {}
    for gene, context in wide.groupby("gene_symbol", sort=False):
        retained = context["passes_SFig6h_display_filter"].all()
        results = set(context["cross_lineage_result"])
        if retained and results == {"shared increase"}:
            cross_contrast[gene] = "shared increase in both mutant contrasts"
        elif retained and results == {"shared decrease"}:
            cross_contrast[gene] = "shared decrease in both mutant contrasts"
        else:
            cross_contrast[gene] = ""
    wide["cross_contrast_result"] = wide["gene_symbol"].map(cross_contrast)

    wide = wide.rename(
        columns={
            "alpha_mean_signed_log2fc": "alpha_mean_fitted_log2FC",
            "alpha_pvalue_l2fc_0_5": "alpha_p_abs_log2FC_gt_0.5",
            "alpha_qvalue_l2fc_0_5": "alpha_BH_q_abs_log2FC_gt_0.5",
            "beta_mean_signed_log2fc": "beta_mean_fitted_log2FC",
            "beta_pvalue_l2fc_0_5": "beta_p_abs_log2FC_gt_0.5",
            "beta_qvalue_l2fc_0_5": "beta_BH_q_abs_log2FC_gt_0.5",
            "original_marker": "curated_marker_in_screen_universe",
        }
    )
    columns = [
        "gene_id",
        "gene_symbol",
        "contrast",
        "num_cells_expressed",
        "BEAM_rank",
        "BEAM_top1000",
        "curated_marker_in_screen_universe",
        "technical_class",
        "alpha_converged",
        "alpha_mean_fitted_log2FC",
        "alpha_p_abs_log2FC_gt_0.5",
        "alpha_BH_q_abs_log2FC_gt_0.5",
        "beta_converged",
        "beta_mean_fitted_log2FC",
        "beta_p_abs_log2FC_gt_0.5",
        "beta_BH_q_abs_log2FC_gt_0.5",
        "max_lineage_BH_q_value",
        "cross_lineage_result",
        "cross_contrast_result",
        "passes_SFig6h_display_filter",
        "SFig6h_category",
        "highlighted_candidate",
    ]
    wide = wide[columns].copy()
    wide["technical_class"] = wide["technical_class"].fillna("")
    contrast_order = pd.Categorical(
        wide["contrast"],
        categories=["dEndo-control", "DKO-control"],
        ordered=True,
    )
    result_order = pd.Categorical(
        wide["cross_lineage_result"],
        categories=[
            "shared increase",
            "shared decrease",
            "not shared at the specified threshold",
        ],
        ordered=True,
    )
    wide = (
        wide.assign(_contrast=contrast_order, _result=result_order)
        .sort_values(
            ["_contrast", "_result", "max_lineage_BH_q_value", "gene_symbol"],
            na_position="last",
        )
        .drop(columns=["_contrast", "_result"])
        .reset_index(drop=True)
    )
    return wide


def plotted_genes() -> pd.DataFrame:
    frame = read_tsv(TABLES / "fig2i_genes.tsv").copy()
    frame["group"] = frame["group"].replace(
        {"alpha output": "alpha terminal output"}
    )
    frame.insert(
        1,
        "lineage",
        frame["group"].str.extract(r"^(alpha|beta)", expand=False),
    )
    frame = frame.rename(
        columns={
            "group": "program",
            "function": "concise_function",
            "selection_rationale": "reason_for_display",
            "beam_top100_rank": "BEAM_top100_rank",
        }
    )
    return frame[[
        "gene",
        "lineage",
        "program",
        "concise_function",
        "reason_for_display",
        "BEAM_top100_rank",
    ]]


def contents(frames: dict[str, pd.DataFrame]) -> pd.DataFrame:
    descriptions = {
        "Samples_QC": (
            "Embryo-level study design, inferred sex, cell retention and "
            "trajectory inclusion. Embryo is the biological replicate."
        ),
        "BEAM": (
            "Complete Monocle2 BEAM result for the alpha/beta branch, including "
            "the genes retained in and labelled on the top-100 heatmap."
        ),
        "Genotype_screen": (
            "Completed 1,116-gene sex-adjusted tradeSeq candidate screen, "
            "restricted to dEndo-control and DKO-control contrasts and shown "
            "once per gene and contrast with alpha and beta effects side by side."
        ),
        "Plotted_genes": (
            "Genes plotted in the lineage-TF and terminal-output comparison. "
            "All 12 genes were identified within the BEAM top 100 and were "
            "then grouped by biological role for display."
        ),
    }
    units = {
        "Samples_QC": "embryo",
        "BEAM": "gene",
        "Genotype_screen": "gene × mutant-control contrast",
        "Plotted_genes": "gene",
    }
    return pd.DataFrame(
        [
            {
                "worksheet": name,
                "rows": len(frame),
                "statistical_or_reporting_unit": units[name],
                "contents": descriptions[name],
            }
            for name, frame in frames.items()
        ]
    )


def style_sheet(worksheet, dataframe: pd.DataFrame, table_name: str) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)
    worksheet.row_dimensions[1].height = 32

    reference = f"A1:{worksheet.cell(worksheet.max_row, worksheet.max_column).coordinate}"
    table = Table(displayName=table_name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    for index, column in enumerate(dataframe.columns, start=1):
        values = [str(column)] + [
            "" if pd.isna(value) else str(value)
            for value in dataframe[column].head(500)
        ]
        width = min(max(max(map(len, values)) + 2, 11), 42)
        if column in {"contents", "concise_function", "reason_for_display"}:
            width = 42
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
        if any(token in column for token in ("p_value", "BH_q", "_q_")):
            for cell in worksheet.iter_cols(
                min_col=index,
                max_col=index,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.number_format = "0.00E+00"
        elif "log2FC" in column:
            for cell in worksheet.iter_cols(
                min_col=index,
                max_col=index,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.number_format = "0.000"


def add_contents_notes(worksheet) -> None:
    start = worksheet.max_row + 3
    worksheet.cell(start, 1, "Statistical notes")
    worksheet.cell(start, 1).fill = SUBHEADER_FILL
    worksheet.cell(start, 1).font = Font(bold=True)
    notes = [
        "Dataset: GSE248369; Davidson et al., Diabetologia (2024), DOI 10.1007/s00125-024-06211-7.",
        "Embryos, not cells, are the biological replicates. The genotype-aware tradeSeq results are candidate-screening statistics and are not embryo-replicated inference.",
        "The genotype screen used six-knot condition-aware NB-GAMs, inferred sex as a covariate, and BH correction separately within each lineage and contrast.",
        "A shared result requires the same signed effect and BH q < 0.05 for the effect-threshold test |log2FC| > 0.5 in both alpha and beta trajectories.",
        "The SFig. 6h display filter additionally required control mean log1p(CP10K) >= 0.1 in both trajectories and finite values for the displayed contrast.",
        "Kcnip4 was the only display-filtered shared increase in both mutant-control contrasts; this is hypothesis-generating rather than mechanistic evidence.",
        "Detailed intermediate tests, smoother coordinates and run manifests remain in the public repository but are intentionally omitted from this workbook.",
    ]
    for offset, note in enumerate(notes, start=1):
        cell = worksheet.cell(start + offset, 1, note)
        worksheet.merge_cells(
            start_row=start + offset,
            start_column=1,
            end_row=start + offset,
            end_column=4,
        )
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.row_dimensions[start + offset].height = 30


def main() -> None:
    beam, beam_ranks = beam_results()
    frames = {
        "Samples_QC": samples_qc(),
        "BEAM": beam,
        "Genotype_screen": genotype_screen(beam_ranks),
        "Plotted_genes": plotted_genes(),
    }
    overview = contents(frames)

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Contents", index=False)
        for name, frame in frames.items():
            frame.to_excel(writer, sheet_name=name, index=False)

        workbook = writer.book
        workbook.properties.title = "Bio-Babel pancreas case-study supplementary tables"
        workbook.properties.subject = "GSE248369 trajectory and genotype-screen results"
        workbook.properties.creator = "Bio-Babel"
        style_sheet(workbook["Contents"], overview, "ContentsIndex")
        add_contents_notes(workbook["Contents"])
        for index, (name, frame) in enumerate(frames.items(), start=1):
            style_sheet(workbook[name], frame, f"SupplementaryData{index}")

    print(f"Wrote {OUTPUT}")
    for name, frame in frames.items():
        print(f"  {name}: {len(frame):,} rows")


if __name__ == "__main__":
    main()
